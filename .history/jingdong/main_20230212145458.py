


from bs4 import BeautifulSoup
import json
import re
import requests
import random
import time
import openpyxl
import os
from openpyxl import Workbook
from urllib.parse import quote


headers={
    "accept": "image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "sec-ch-ua": "\"Chromium\";v=\"110\", \"Not A(Brand\";v=\"24\", \"Microsoft Edge\";v=\"110\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "image",
    "sec-fetch-mode": "no-cors",
    "sec-fetch-site": "same-site",
    "cookie": "shshshfpa=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; shshshfpb=yZaSjZXfeZ-R-j6eY1Nun6A; cud=bb70d57435d5263f9fc904a6ba2ad5b8; __jdu=16722123231461165655386; areaId=15; ipLoc-djd=15-1213-3038-59931; pinId=F0lqqKsS7Ce1ffsfn98I-w; pin=LOSER%E4%B8%B0; unick=LOSER%E4%B8%B0; _tp=Iz8Q38tsk7Xe1Xc1Cy0l4Q%3D%3D; _pst=LOSER%E4%B8%B0; TrackID=153eT-bS7-TH0zTiURvbcsXJ906H2je9nODkD5QrvdAOSH3lh5WOMlE_E_WvGLCO6VwC1LDNjV1QzURU_Em52NPCWhCOon3yKMb1h5OHxvK_zNdqe0N90kh5n5LHhz2Na; ceshi3.com=000; unpl=JF8EAK1nNSttD09UBU9WSBZEGVhcW1xdGEcDPGVXVVReHFIMTAUeGxN7XlVdXxRKFB9vZBRUX1NPUA4fACsSEHteVVxdAE4RAmpvNWRYW0xTBCsBGyIRe11TWF4LSxQBbGUFUVpYS1MBHAYfEBlMbVVuXQBMJwNuZgRRXV9LVQIZAysiEXtcZF9tXiUXAm5mBF1aWE1USBsFHRETS15WXV8IThADb2ABU1lcSV0CKwMrEQ; __jdv=122270672|kong|t_2031947049_|jingfen|f511eeb5ec4940dba1b3c087f78f6482|1676181354851; cvt=10; __jda=76161171.16722123231461165655386.1672212323.1676180315.1676184513.18; __jdc=76161171; PCSYCityID=CN_330000_330100_0; user-key=4cac354c-fa26-49a7-a778-7f31be11b349; cn=1; thor=B0FCAB6E82CA81C65C8CD0C89AFF0FD42C09EC761E24C53E3B9B787FDEA9C34FD4B887E696A2E50CE80DBD9F72C833DE9215DC5FE8A62B34F63BB2AF314817CB35B296A7AC2568AE387E24E696A96F12A7B54D478DAC98479095B421A0DBF67518A10244A0596CE20B922C4B0BEDD89AB2A68BD246428EEE8657E300A34FD0F9; __jdb=76161171.7.16722123231461165655386|18.1676184513; csn=7; shshshfp=bc1f4aeb46f0dda1cb207282332d7d4a; shshshfpx=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; shshshsID=5bc7586f74bcc7a33c6e1a4097d29f49_6_1676184649964; 3AB9D23F7A4B3C9B=FTSDKAYZNZKCZEI2SHBQY2UVW7R5ZBD7C4KZOV6I52DX4YPOAOJ2OLORCGQB5P35WXNCTLBOXG6CZWIZCSDSGUAJ5U",
    "Referrer-Policy": "no-referrer-when-downgrade",
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.78"
}

SEARCH_KEYWORD='正大食品（CP）京东自营旗舰店'

TOTAL_PAGE=6  #这里懒得自动了，直接手动算了
PAGE_SIZE=3


START_IDX=96
DATA={}



def product_spider():
    product_id_list=[]
    product_price_list=[]
    product_name_list=[]
    product_shop_list=[]
    
    #search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&enc=utf-8'


    # for i in range(0,TOTAL_PAGE):
    #     time.sleep(random.randint(3,4))
    #     cur_page=i*2+1
        
    #     page=get_search_page(cur_page)

    #     file=open('./tmp/'+str(cur_page)+'.txt','wb')
    #     file.write(page)
    #     file.close()
    
    for i in range(0,TOTAL_PAGE):
        time.sleep(random.randint(3,4))
        cur_page=i+1
        page=get_search_page(cur_page)

        #print(page)


        product_id_list.extend(re.findall(r'<li data-sku="([0-9]*)"',page,re.S))
        product_price_list.extend(re.findall(r'<i data-price="[0-9]*">([0-9.]*)</i>',page,re.S))
        product_name_list.extend(re.findall(r'<div class="p-name p-name-type-2">.*?<a target="_blank" title="(.*?)"',page,re.S)) 
        product_shop_list.extend(re.findall(r'<a target="_blank" class="curr-shop hd-shopname" .*?>(.*?)</a>',page,re.S))
        #product_id_list.extend(re.findall())


    # print(product_name_list)
    # print(product_price_list)
    # print(product_shop_list)
    print(f'''product_id_list:{len(product_id_list)},\n
product_price_list:{len(product_price_list)},\n
product_name_list:{len(product_name_list)},\n
product_shop_list:{len(product_shop_list)}\n   
        ''')
    wb=Workbook()
    ws=wb.active
    ws.append(['product_id_list','product_name_list','product_price_list','product_shop_list'])

    for i in range(0,len(product_id_list)):
        ws.append([product_id_list[i],product_name_list[i],product_price_list[i],product_shop_list[i]])

    wb.save('商品信息.xlsx')

    DATA['product_id_list']=product_id_list
    DATA['product_name_list']=product_name_list
    DATA['product_price_list']=product_price_list
    DATA['product_shop_list']=product_shop_list


    with open('DATA.json', 'w') as file_obj:
        json.dump(DATA,file_obj)
    



def comment_spider(product_id):
    
    wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')


    ws=wb.create_sheet(f'{product_id}_comment')
    
    ws.append(['creationTime','nickname','score','content'])
    for cur_page in range(0,99):
        #time.sleep(random.randint(2,3))
        url=f'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId={product_id}&score=0&sortType=5&page={cur_page}&pageSize=10&isShadowSku=0&fold=1'
        
        
        
        res=requests.get(url=url,headers=headers)
        #print(res.text)
        if(res.text==''):
            break
        

        data_json=re.match('fetchJSON_comment98\((.*)\);',res.text).group(1)
       # print(data_json)
        #time.sleep(10)

        jsonObj=json.loads(data_json)
        for comment in jsonObj["comments"]:
            ws.append([comment["creationTime"],comment["nickname"],comment["score"],comment["content"]])
        
    wb.save(os.getcwd() + '/test.xlsx')


def get_search_page(cur_page):
    search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&page={cur_page}&qrst=1&psort=5&click=1'
    res=requests.get(url=search_url,headers=headers)
    #html = BeautifulSoup(res.text, 'lxml')
    return res.text


def main():
    # print('正在爬取商品信息')
    # product_spider()

    # print('开始爬取评论')

    # wb=Workbook()
    # wb.save(os.getcwd() + '/test.xlsx')

    with open('DATA.json') as file_obj:
        DATA=json.load(file_obj)
    total=len(DATA["product_id_list"])
    print(f'共{total}个商品')
    idx=START_IDX

    for idx in range(START_IDX,total):
        id=DATA['product_id_list'][idx]
        print(f'正在爬取{id}的评论-------{idx}/{total}')
        comment_spider(id)
        print(f'爬取完毕')
        time.sleep(random.randint(2,3))
    
 







if __name__=='__main__':
    main()


