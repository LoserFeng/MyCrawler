


from bs4 import BeautifulSoup
import json
import re
import requests
import random
import time
import openpyxl
import os
from openpyxl import Workbook
from urllib.parse import quote


headers={
       "accept": "image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "sec-ch-ua": "\"Chromium\";v=\"110\", \"Not A(Brand\";v=\"24\", \"Microsoft Edge\";v=\"110\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "image",
    "sec-fetch-mode": "no-cors",
    "sec-fetch-site": "same-site",
    "cookie": "shshshfpb=yZaSjZXfeZ-R-j6eY1Nun6A; shshshfpa=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; areaId=15; ipLoc-djd=15-1213-3038-59931; TrackID=1urjIeDGRd74BYvrLyY3Dc2b2Kyj-lLWGdpxThqlzMVFVBO8QACEoShYm6wv26GSV1r_WY1t7MtbBd13Mt3rgBPPbwSvZtVq7800KGPEEuZDzofFBSf62hbKLxrkJ40f-; pinId=F0lqqKsS7Ce1ffsfn98I-w; pin=LOSER%E4%B8%B0; unick=LOSER%E4%B8%B0; ceshi3.com=000; _tp=Iz8Q38tsk7Xe1Xc1Cy0l4Q%3D%3D; _pst=LOSER%E4%B8%B0; __jdu=1676185295741330850847; PCSYCityID=CN_330000_330100_0; shshshfp=f3a35f8c7bd61322ba4dc0413d247ffb; shshshfpx=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; token=bd1fb4e4df3fff104d28b636581b3a4a,3,931214; __tk=RUP0VUbyVUVyVZaNi0eMRZjEVDeFRxyNWlk0WliJWUp0RcV0VZy5RX,3,931214; jsavif=0; __jdc=122270672; thor=B0FCAB6E82CA81C65C8CD0C89AFF0FD4946E83CA3DE1E9E9E28B77099DC7D5A0CA0727C33F8BA179AB1D8C8C4E42B2FF2609621A6A855FBDEA47F3DD799E029C0D090C0578D36F2944CF18C96EC668658C7594634811BA2B8F7844E1613329F2F2D72876E5AD113DD3F0E0B95F537347F524319AFB1F48E6252ABED53ABF44BC; ip_cityCode=1213; unpl=JF8EAKtnNSttD0NcURNWExUSH1QGW1sLQx4GOG8FVAhcSQBRGFYfGxd7XlVdXxRKFB9vZBRUWVNKXQ4ZASsSEHteVVxdAE4RAmpvNWRYW0xTBCsBGyIRe11TWFoIQhMAZ2MDUV1aSVUMHwMSFxZ7XGReVQ97FwJuZgRTWFlJXQcTMisTIEptVW4LZksWAm5mDFNdXksZBRwEHBIZT15cWlsNSxUBbm4BVVRdTWQEKwE; __jda=122270672.1676185295741330850847.1676185296.1676185309.1676185714.3; __jdv=122270672|kong|t_2031947049_|jingfen|f99e9e963e8c472984f911d53ee2e586|1676185713804; shshshsID=65aebb41e16cb75563af1e0514703f3c_11_1676185745959; __jdb=122270672.2.1676185295741330850847|3.1676185714; 3AB9D23F7A4B3C9B=C435LG3UEXGMHAEILE7G3HDVXAXNBB2CCFDLR6B4JYHHDL7YQD7DBXWW3GDPN6URM5RVNTAQMJK77BOJ7SZ6MWLZXY",
    "Referer": "https://item.jd.com/",
    "Referrer-Policy": "strict-origin-when-cross-origin",
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.78"
}

SEARCH_KEYWORD='正大食品（CP）京东自营旗舰店'

TOTAL_PAGE=6  #这里懒得自动了，直接手动算了
PAGE_SIZE=3


START_IDX=96
DATA={}



def product_spider():
    product_id_list=[]
    product_price_list=[]
    product_name_list=[]
    product_shop_list=[]
    
    #search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&enc=utf-8'


    # for i in range(0,TOTAL_PAGE):
    #     time.sleep(random.randint(3,4))
    #     cur_page=i*2+1
        
    #     page=get_search_page(cur_page)

    #     file=open('./tmp/'+str(cur_page)+'.txt','wb')
    #     file.write(page)
    #     file.close()
    
    for i in range(0,TOTAL_PAGE):
        time.sleep(random.randint(3,4))
        cur_page=i+1
        page=get_search_page(cur_page)

        #print(page)


        product_id_list.extend(re.findall(r'<li data-sku="([0-9]*)"',page,re.S))
        product_price_list.extend(re.findall(r'<i data-price="[0-9]*">([0-9.]*)</i>',page,re.S))
        product_name_list.extend(re.findall(r'<div class="p-name p-name-type-2">.*?<a target="_blank" title="(.*?)"',page,re.S)) 
        product_shop_list.extend(re.findall(r'<a target="_blank" class="curr-shop hd-shopname" .*?>(.*?)</a>',page,re.S))
        #product_id_list.extend(re.findall())


    # print(product_name_list)
    # print(product_price_list)
    # print(product_shop_list)
    print(f'''product_id_list:{len(product_id_list)},\n
product_price_list:{len(product_price_list)},\n
product_name_list:{len(product_name_list)},\n
product_shop_list:{len(product_shop_list)}\n   
        ''')
    wb=Workbook()
    ws=wb.active
    ws.append(['product_id_list','product_name_list','product_price_list','product_shop_list'])

    for i in range(0,len(product_id_list)):
        ws.append([product_id_list[i],product_name_list[i],product_price_list[i],product_shop_list[i]])

    wb.save('商品信息.xlsx')

    DATA['product_id_list']=product_id_list
    DATA['product_name_list']=product_name_list
    DATA['product_price_list']=product_price_list
    DATA['product_shop_list']=product_shop_list


    with open('DATA.json', 'w') as file_obj:
        json.dump(DATA,file_obj)
    



def comment_spider(product_id):
    
    wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')


    ws=wb.create_sheet(f'{product_id}_comment')
    
    ws.append(['creationTime','nickname','score','content'])
    for cur_page in range(0,99):
        #time.sleep(random.randint(2,3))
        url=f'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId={product_id}&score=0&sortType=5&page={cur_page}&pageSize=10&isShadowSku=0&fold=1'
        
        
        
        res=requests.get(url=url,headers=headers)
        #print(res.text)
        if(res.text==''):
            print('该商品没有评论')
            return False
        

        data_json=re.match('fetchJSON_comment98\((.*)\);',res.text).group(1)
       # print(data_json)
        #time.sleep(10)

        jsonObj=json.loads(data_json)
        for comment in jsonObj["comments"]:
            ws.append([comment["creationTime"],comment["nickname"],comment["score"],comment["content"]])
        
    wb.save(os.getcwd() + '/test.xlsx')
    return True


def get_search_page(cur_page):
    search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&page={cur_page}&qrst=1&psort=5&click=1'
    res=requests.get(url=search_url,headers=headers)
    #html = BeautifulSoup(res.text, 'lxml')
    return res.text


def main():
    # print('正在爬取商品信息')
    # product_spider()

    # print('开始爬取评论')

    # wb=Workbook()
    # wb.save(os.getcwd() + '/test.xlsx')

    with open('DATA.json') as file_obj:
        DATA=json.load(file_obj)
    total=len(DATA["product_id_list"])
    print(f'共{total}个商品')
    idx=START_IDX

    for idx in range(START_IDX,total):
        id=DATA['product_id_list'][idx]
        print(f'正在爬取{id}的评论-------{idx}/{total}')
        flag=comment_spider(id)
        if not flag:
            print('爬虫终止')
            return -1
        print(f'爬取完毕')
        time.sleep(random.randint(2,3))
    
 







if __name__=='__main__':
    main()


