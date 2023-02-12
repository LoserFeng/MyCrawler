


from bs4 import BeautifulSoup
import json
import re
import requests
import random
import time
import openpyxl
import os
from openpyxl import Workbook
from urllib.parse import quote


headers={


{
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "cache-control": "max-age=0",
    "sec-ch-ua": "\"Chromium\";v=\"110\", \"Not A(Brand\";v=\"24\", \"Microsoft Edge\";v=\"110\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "none",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "cookie":'shshshfpa=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; shshshfpb=yZaSjZXfeZ-R-j6eY1Nun6A; __jdu=16722123231461165655386; o2State={"webp":true,"avif":false}; areaId=15; PCSYCityID=CN_330000_330100_0; ipLoc-djd=15-1213-3038-59931; pinId=F0lqqKsS7Ce1ffsfn98I-w; pin=LOSER丰; unick=LOSER丰; _tp=Iz8Q38tsk7Xe1Xc1Cy0l4Q==; _pst=LOSER丰; unpl=JF8EAKhnNSttXE5dURkLHxYVQwhVW11bHkdQPTUHAQ0NTwNXTgYeQUB7XlVdXxRKFB9vZBRUX1NLUA4fAysSEHtdVV9fCU0eAGhgNWRYW0xTBCsBGyIRe11TWF4LSxQBbGUFUVpYS1MBHAYfEBlMbVVuXQBMJwRvZgJVX1h7ZAQrAysTIB0zVF9cCUofBWxkBhldX01XBhsBGRESS1hTXl0PTxAHa2UMU21Ze1c; __jdv=122270672|kong|t_1003078266_|jingfen|548e385549d041bdabcc3dad5fcd54ba|1676178951001; TrackID=153eT-bS7-TH0zTiURvbcsXJ906H2je9nODkD5QrvdAOSH3lh5WOMlE_E_WvGLCO6VwC1LDNjV1QzURU_Em52NPCWhCOon3yKMb1h5OHxvK_zNdqe0N90kh5n5LHhz2Na; ceshi3.com=000; shshshfpx=a59ef60d-c8a6-92ba-2baf-a646895cc98d-1672212324; jsavif=0; __jda=122270672.16722123231461165655386.1672212323.1676178924.1676178951.16; __jdc=122270672; shshshfp=f3a35f8c7bd61322ba4dc0413d247ffb; thor=B0FCAB6E82CA81C65C8CD0C89AFF0FD42C09EC761E24C53E3B9B787FDEA9C34FF627D6AEFC38BCAC4D9E55B95A6F46D61FC30F474445D8422545333C3F7D288901B7939CCE16CD1A3108C8105A17C0CE28B666A847F3C373CFE3641353FF1B0969ACDC24CEA29B24A6C086B57A2E6B033526E499C031623582A7BF4C9AF7C1A9; shshshsID=644c61c7f485dcd2b6d7bdff14f3eb4f_14_1676179764040; joyya=0.1676179768.23.1d99xdw; 3AB9D23F7A4B3C9B=C435LG3UEXGMHAEILE7G3HDVXAXNBB2CCFDLR6B4JYHHDL7YQD7DBXWW3GDPN6URM5RVNTAQMJK77BOJ7SZ6MWLZXY; __jdb=122270672.19.16722123231461165655386|16.1676178951',
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.78"
}

SEARCH_KEYWORD='正大食品（CP）京东自营旗舰店'

TOTAL_PAGE=6  #这里懒得自动了，直接手动算了
PAGE_SIZE=3


START_IDX=0
DATA={}



def product_spider():
    product_id_list=[]
    product_price_list=[]
    product_name_list=[]
    product_shop_list=[]
    
    #search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&enc=utf-8'


    # for i in range(0,TOTAL_PAGE):
    #     time.sleep(random.randint(3,4))
    #     cur_page=i*2+1
        
    #     page=get_search_page(cur_page)

    #     file=open('./tmp/'+str(cur_page)+'.txt','wb')
    #     file.write(page)
    #     file.close()
    
    for i in range(0,TOTAL_PAGE):
        time.sleep(random.randint(3,4))
        cur_page=i+1
        page=get_search_page(cur_page)

        #print(page)


        product_id_list.extend(re.findall(r'<li data-sku="([0-9]*)"',page,re.S))
        product_price_list.extend(re.findall(r'<i data-price="[0-9]*">([0-9.]*)</i>',page,re.S))
        product_name_list.extend(re.findall(r'<div class="p-name p-name-type-2">.*?<a target="_blank" title="(.*?)"',page,re.S)) 
        product_shop_list.extend(re.findall(r'<a target="_blank" class="curr-shop hd-shopname" .*?>(.*?)</a>',page,re.S))
        #product_id_list.extend(re.findall())


    # print(product_name_list)
    # print(product_price_list)
    # print(product_shop_list)
    print(f'''product_id_list:{len(product_id_list)},\n
product_price_list:{len(product_price_list)},\n
product_name_list:{len(product_name_list)},\n
product_shop_list:{len(product_shop_list)}\n   
        ''')
    wb=Workbook()
    ws=wb.active
    ws.append(['product_id_list','product_name_list','product_price_list','product_shop_list'])

    for i in range(0,len(product_id_list)):
        ws.append([product_id_list[i],product_name_list[i],product_price_list[i],product_shop_list[i]])

    wb.save('商品信息.xlsx')

    DATA['product_id_list']=product_id_list
    DATA['product_name_list']=product_name_list
    DATA['product_price_list']=product_price_list
    DATA['product_shop_list']=product_shop_list


    with open('DATA.json', 'w') as file_obj:
        json.dump(DATA,file_obj)
    



def comment_spider(product_id):
    
    wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')


    ws=wb.create_sheet(f'{product_id}_comment')
    
    ws.append(['creationTime','nickname','score','content'])
    for cur_page in range(0,99):
        time.sleep(random.randint(2,3))
        url=f'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId={product_id}&score=0&sortType=5&page={cur_page}&pageSize=10&isShadowSku=0&fold=1'
        res=requests.get(url=url,headers=headers)
        print(res.text)
        

        data_json=re.match('fetchJSON_comment98\((.*)\);',res.text).group(1)
       # print(data_json)
        #time.sleep(10)

        jsonObj=json.loads(data_json)
        for comment in jsonObj["comments"]:
            ws.append([comment["creationTime"],comment["nickname"],comment["score"],comment["content"]])
        
    wb.save(os.getcwd() + '/test.xlsx')


def get_search_page(cur_page):
    search_url=f'https://search.jd.com/Search?keyword={quote(SEARCH_KEYWORD)}&page={cur_page}&qrst=1&psort=5&click=1'
    res=requests.get(url=search_url,headers=headers)
    #html = BeautifulSoup(res.text, 'lxml')
    return res.text


def main():
    # print('正在爬取商品信息')
    # product_spider()

    # print('开始爬取评论')

    # wb=Workbook()
    # wb.save(os.getcwd() + '/test.xlsx')

    with open('DATA.json') as file_obj:
        DATA=json.load(file_obj)
    total=len(DATA["product_id_list"])
    print(f'共{total}个商品')
    idx=0
    for id in DATA['product_id_list']:
        print(f'正在爬取{id}的评论-------{idx}/{total}')
        comment_spider(id)
        print(f'爬取完毕')
        idx+=1
        time.sleep(random.randint(2,3))
    
 







if __name__=='__main__':
    main()


