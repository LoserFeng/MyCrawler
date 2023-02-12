from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import openpyxl
import os


wb=Workbook()

ws=wb.create_sheet('hello')


ws.append(['asd'])


wb.save(os.getcwd() +'/demo1.xlsx')



for i in range(1,100):
    wb=openpyxl.load_workbook('./demo1.xlsx')
    wb.create_sheet(f'hello{i}')


    wb.save(os.getcwd() +'/demo1.xlsx') 




