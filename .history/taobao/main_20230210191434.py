

import requests
import time
import json
import spider_utils
import random
import os
import requests
import re
import sys
import openpyxl
import logging
from openpyxl import Workbook
from selenium.webdriver.common.by import By
import threading
import json



#基础数据

login_id='15157165006'
login_password='110900fyf'

KEY_WORD='正大食品旗舰店'

DATA_FILE_NAME='DATA.json'





def init_log():
    logging.basicConfig(level=logging.INFO,
                    filename='./log.txt',
                    filemode='w',
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')
    # logging.info('这是 loggging info message')
    # logging.debug('这是 loggging debug message')
    # logging.warning('这是 loggging a warning message')
    # logging.error('这是 an loggging error message')
    # logging.critical('这是 loggging critical message')







def main():
    product_id_list=[]
    print("where")
    init_log()
    wb=Workbook()
    ws=wb.active  #worksheet
    ws.append(['product_id', 'raw_title', 'view_price', 'view_sales', 'comment_count'
    , 'area', 'userid','name','detail_url', "categary"])
    
    driver=spider_utils.get_browser()
    driver.implicitly_wait(60)
    logging.info('初始化完成')
    

    spider_utils.login(driver,login_id,login_password)
    logging.info('完成登录')
    time.sleep(random.randint(3, 5))

    t1 = threading.Thread(target=spider_utils.monitor, args=(driver,))
    t1.start()
    logging.info('已经启动monitor')
    flag=spider_utils.verify(driver)
    if not flag:
        return -1
    #url = f'https://s.taobao.com/search?q={quote(KEYWORD)}'
    url = f'https://s.taobao.com/search?q={KEY_WORD}'
    driver.get(url)
    time.sleep(random.randint(3,5))

    while driver.title != f'{KEY_WORD}_淘宝搜索':
        logging.info('不是正确的页面，等待!')
        time.sleep(5)

    # flag=spider_utils.verify(driver)
    # if not flag:
    #     return -1
    
    page_str = driver.find_element(By.CSS_SELECTOR, "div.total").text
    total_page=int(page_str.split(" ")[1])

    logging.info(f'总共搜索到 {total_page}页')

    logging.info('开始爬虫读取数据')
    for page in range(0,total_page):
        print('page: ',page+1)
        logging.info(f'正在爬取第{page+1}页')
        #url = f'https://s.taobao.com/search?q={quote(KEYWORD)}&s={48 * page}'
        url = f'https://s.taobao.com/search?q={KEY_WORD}&s={44 * page}'
        time.sleep(random.randint(3, 5))
        driver.get(url)
        time.sleep(random.randint(2, 4))

        # flag=spider_utils.verify(driver)
        # if not flag:
        #     return -2
        while driver.title != f'{KEY_WORD}_淘宝搜索':
            logging.info('不是正确的页面，等待!')
            time.sleep(5)
        html_text = driver.page_source
        try:
            json_str = re.findall(r'g_page_config = (.*?) g_srp_loadCss', html_text, re.S)[0].strip()[:-1]
        except:
            print("NG", url)
            continue
        # 格式化
        json_dict = json.loads(json_str)

        # 获取信息列表
        auctions = json_dict['mods']['itemlist']['data']['auctions']
        DATA = []
        # 提取数据 注意是否有通过异步进行加载的数据，需要在进行请求
        for auction in auctions:
            temp = {
                'product_id': '' if 'nid' not in auction else auction['nid'],
                'raw_title': '' if 'raw_title' not in auction else auction['raw_title'],
                'view_price': '' if 'view_price' not in auction else auction['view_price'],
                'view_sales': '' if 'view_sales' not in auction else auction['view_sales'],
                'comment_count': '' if 'comment_count' not in auction else auction['comment_count'],
                # 'view_fee': '否' if float(auction['view_fee']) else '是',
                # 'isTmall': '是' if auction['shopcard']['isTmall'] else '否',
                'area': '' if 'item_loc' not in auction else auction['item_loc'],
                'userid': '' if 'user_id' not in auction else auction['user_id'],
                'name': '' if 'nick' not in auction else auction['nick'],
                'detail_url': '' if 'detail_url' not in auction else auction['detail_url'],
            }
            DATA.append(temp)
            product_id_list.append(temp['product_id'])
            ws.append([temp['product_id'], temp['raw_title'], temp['view_price'], temp['view_sales'],
                        temp['comment_count'],
                        temp['area'], temp['userid'], temp['name'], temp['detail_url'], KEY_WORD])
        # 覆盖方式保存表格

        wb.save(os.getcwd() + '/test.xlsx')
        wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')
        # 对xlsx文件操作之前,要 active.
        ws = wb.active
    print(DATA)


    
    with open(DATA_FILE_NAME, 'w') as file_obj:
        json.dump(product_id_list,file_obj)





    wb.save(os.getcwd() + '/test.xlsx')
    logging.info('保存商品信息')


    logging.info('开始爬取评论')
    wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')
    with open(DATA_FILE_NAME) as file_obj:
        DATA=json.load(file_obj)
    

    for product_id in product_id_list:
        ws=wb.create_sheet('hello')
        logging.info(f'正在爬取ID为{product_id}的评论')
        url = f'https://detail.tmall.com/item.htm?id={product_id}'
        time.sleep(random.randint(3, 5))
        driver.get(url)
        time.sleep(random.randint(2, 4))


        #comment_btn=driver.find_element(By.XPATH,value='//*[@id="root"]/div/div[2]/div[2]/div[2]/div[1]/div/div/div[2]/span')
        comment_btn=driver.find_element(By.XPATH,value="//*[text()='宝贝评价']")
        comment_btn.click()
        time.sleep(random.randint(2, 4))


        



        while driver.title == '验证码拦截':
            logging.info('不是正确的页面，等待!')
            time.sleep(5)
        html_text = driver.page_source
        try:
            json_str = re.findall(r'g_page_config = (.*?) g_srp_loadCss', html_text, re.S)[0].strip()[:-1]
        except:
            print("NG", url)
            continue
        # 格式化
        json_dict = json.loads(json_str)

        # 获取信息列表
        auctions = json_dict['mods']['itemlist']['data']['auctions']
        DATA = []
            





    
    spider_utils.stop_monitor()
    t1.join()
    driver.quit()
    exit()









if __name__=='__main__':
    main()









