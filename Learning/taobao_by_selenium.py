# !/user/bin/env python
# -*- coding: utf-8 -*-
import os
import time
import re
import json
import random
import requests
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
# 实现规避检测
from selenium.webdriver import ChromeOptions
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from urllib.parse import quote

# 淘宝评论 https://blog.csdn.net/qq_36333776/article/details/121597417
# 滑块破解 https://blog.csdn.net/z__Java__/article/details/110749144


# 输入你的淘宝账号密码
login_id = '15157165006'
login_password = '110900fyf'

# 输入搜索关键字
keyword_list = ['车衣', '纸巾']


# 设置为开发者模式，防止被各大网站识别出来使用了Selenium console中输入window.navigator.webdriver 测试
def get_browser():
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument('--ignore-certificate-errors')   #主要是该条
    options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome(options=options)
    # 使用js关闭检测机制
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                        Object.defineProperty(navigator, 'webdriver', {
                          get: () => undefined
                        })
                      """
    })
    return driver


def login(login_id, login_password):
    driver.get('https://login.taobao.com/')
    print('title is ', driver.title)
    # if driver.title=='验证码拦截':
    #     slider(url)

    # 发送命令-查找元素与操作元素
    # 找到输入账号框，清除框内信息，再输入你的账号
    driver.find_element(By.CSS_SELECTOR, '#fm-login-id').clear()
    driver.find_element(By.CSS_SELECTOR, '#fm-login-id').send_keys(login_id)
    # 找到输入密码框，清除框内信息，再输入你的密码
    driver.find_element(By.CSS_SELECTOR, '#fm-login-password').clear()
    driver.find_element(By.CSS_SELECTOR, '#fm-login-password').send_keys(login_password + Keys.ENTER)


def slider(url):
    print(driver.title, url)
    # if driver.title == '验证码拦截':
    #     print(driver.title, url)
    #     time.sleep(random.randint(3, 5))  # 随机休眠5-10s
    #     try:
    #         if driver.find_element(By.CSS_SELECTOR, "#\`nc_1_refresh1\`").is_displayed():
    #             driver.find_element(By.CSS_SELECTOR, "#\`nc_1_refresh1\`").click()
    #             time.sleep(random.randint(1, 3))
    #             slider = driver.find_element(By.CSS_SELECTOR, '#nc_1_n1z')
    #
    #     except:
    #         slider = driver.find_element(By.CSS_SELECTOR, '#nc_1_n1z')
    #         pass

    slider = driver.find_element(By.CSS_SELECTOR, '#nc_1_n1z')
    # 拖拽滑块
    action = ActionChains(driver)
    action.click_and_hold(slider)
    sum = 0
    while True:
        x = random.randint(5, 70)
        action.move_by_offset(x, 0)
        time.sleep((random.randint(1, 2)) / 10)
        sum += x
        if sum >= 260:
            break
    time.sleep(1)
    action.release().perform()
    time.sleep(1)


def data_analysis(url, KEYWORD):
    time.sleep(random.randint(3, 5))
    driver.get(url)
    time.sleep(random.randint(3, 5))
    html_text = driver.page_source
    try:
        json_str = re.findall(r'g_page_config = (.*?) g_srp_loadCss', html_text, re.S)[0].strip()[:-1]
    except:
        print("NG", url)
    # 格式化
    json_dict = json.loads(json_str)

    # 获取信息列表
    auctions = json_dict['mods']['itemlist']['data']['auctions']
    DATA = []
    # 提取数据 注意是否有通过异步进行加载的数据，需要在进行请求
    for auction in auctions:
        temp = {
            'product_id': '' if 'nid' not in auction else auction['nid'],
            'raw_title': '' if 'raw_title' not in auction else auction['raw_title'],
            'view_price': '' if 'view_price' not in auction else auction['view_price'],
            'view_sales': '' if 'view_sales' not in auction else auction['view_sales'],
            'comment_count': '' if 'comment_count' not in auction else auction['comment_count'],
            # 'view_fee': '否' if float(auction['view_fee']) else '是',
            # 'isTmall': '是' if auction['shopcard']['isTmall'] else '否',
            'area': '' if 'item_loc' not in auction else auction['item_loc'],
            'userid': '' if 'user_id' not in auction else auction['user_id'],
            'name': '' if 'nick' not in auction else auction['nick'],
            'detail_url': '' if 'detail_url' not in auction else auction['detail_url'],
        }
        DATA.append(temp)
        wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')
        # 对xlsx文件操作之前,要 active.
        ws = wb.active
        # 覆盖方式保存表格
        ws.append([temp['product_id'], temp['raw_title'], temp['view_price'], temp['view_sales'],
                   temp['comment_count'],
                   temp['area'], temp['userid'], temp['name'], temp['detail_url'], KEYWORD])
    wb.save(os.getcwd() + '/test.xlsx')


# 创建表格
wb = Workbook()
ws = wb.active
# 写入一行的数据, 从第一列的空白单元格开始
ws.append(
    ['product_id', 'raw_title', 'view_price', 'view_sales', 'comment_count', 'area', 'userid',
     'name', 'detail_url', "categary"])

# 登录
driver = get_browser()
driver.implicitly_wait(60)
login(login_id, login_password)
time.sleep(random.randint(3, 5))

for KEYWORD in keyword_list:
    i = 1
    while driver.title == '验证码拦截':
        i = i + 1
        time.sleep(random.randint(2, 3))
        slider('logining')
        if i >= 2:
            # 打开新标签页并切换到新标签页
            driver.refresh()
            # driver.switch_to.new_window('tab')

        time.sleep(random.randint(3, 5))

    url = f'https://s.taobao.com/search?q={quote(KEYWORD)}'
    driver.get(url)
    time.sleep(random.randint(3, 5))

    i = 1
    while driver.title == '验证码拦截':
        i = i + 1
        time.sleep(random.randint(2, 3))
        slider('logining')

        if i >= 2:
            # 打开新标签页并切换到新标签页
            driver.refresh()
            # driver.switch_to.new_window('tab')

        time.sleep(random.randint(3, 5))
    page_str = driver.find_element(By.CSS_SELECTOR, "div.total").text
    total_page = int(page_str.split(" ")[1])
    # mainsrp-pager > div > div > div > div.total
    # total_page=driver.get(f'https://s.taobao.com/search?q={quote(KEYWORD)})
    # total_page = 100

    for page in range(0, total_page):
        print("page: ", page + 1)
        url = f'https://s.taobao.com/search?q={quote(KEYWORD)}&s={48 * page}'
        time.sleep(random.randint(3, 5))
        driver.get(url)
        time.sleep(random.randint(2, 3))

        i = 1
        while driver.title == '验证码拦截':
            i = i + 1
            time.sleep(random.randint(2, 3))
            slider(url)
            if i >= 2:
                # 打开新标签页并切换到新标签页
                driver.refresh()
                # driver.switch_to.new_window('tab')

            time.sleep(random.randint(4, 7))
            if driver.title != '验证码拦截':
                data_analysis(url, KEYWORD)

        else:
            html_text = driver.page_source
            try:
                json_str = re.findall(r'g_page_config = (.*?) g_srp_loadCss', html_text, re.S)[0].strip()[:-1]
            except:
                print("NG", url)
                continue
            # 格式化
            json_dict = json.loads(json_str)

            # 获取信息列表
            auctions = json_dict['mods']['itemlist']['data']['auctions']
            DATA = []
            # 提取数据 注意是否有通过异步进行加载的数据，需要在进行请求
            for auction in auctions:
                temp = {
                    'product_id': '' if 'nid' not in auction else auction['nid'],
                    'raw_title': '' if 'raw_title' not in auction else auction['raw_title'],
                    'view_price': '' if 'view_price' not in auction else auction['view_price'],
                    'view_sales': '' if 'view_sales' not in auction else auction['view_sales'],
                    'comment_count': '' if 'comment_count' not in auction else auction['comment_count'],
                    # 'view_fee': '否' if float(auction['view_fee']) else '是',
                    # 'isTmall': '是' if auction['shopcard']['isTmall'] else '否',
                    'area': '' if 'item_loc' not in auction else auction['item_loc'],
                    'userid': '' if 'user_id' not in auction else auction['user_id'],
                    'name': '' if 'nick' not in auction else auction['nick'],
                    'detail_url': '' if 'detail_url' not in auction else auction['detail_url'],
                }
                DATA.append(temp)
                ws.append([temp['product_id'], temp['raw_title'], temp['view_price'], temp['view_sales'],
                           temp['comment_count'],
                           temp['area'], temp['userid'], temp['name'], temp['detail_url'], KEYWORD])
            # 覆盖方式保存表格
            wb.save(os.getcwd() + '/test.xlsx')
            wb = openpyxl.load_workbook(os.getcwd() + '/test.xlsx')
            # 对xlsx文件操作之前,要 active.
            ws = wb.active

wb.save(os.getcwd() + '/test.xlsx')
driver.quit()
